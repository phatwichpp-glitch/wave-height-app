"""
core/export.py
PDF (ReportLab) and Excel (openpyxl) export generators.
"""

import io
import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from typing import List, Dict, Any, Optional

# ─── Excel ────────────────────────────────────────────────────────────────────

def build_excel(results: List[Dict[str, Any]], selected_epoch_idx: int = 0) -> bytes:
    """
    Build an Excel workbook with 3 sheets:
      Sheet 1 "สรุป"         — mean/max/min/std for all 3 Hs methods
      Sheet 2 "อนุกรมเวลา"   — full epoch table
      Sheet 3 "สเปกตรัม"     — f and Pxx arrays for selected epoch
    Returns bytes.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

    def style_header_row(ws, row_num: int):
        for cell in ws[row_num]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "สรุป"

    summary_headers = ["วิธีการ", "ค่าเฉลี่ย (m)", "ค่าสูงสุด (m)",
                        "ค่าต่ำสุด (m)", "ค่าเบี่ยงเบนมาตรฐาน (m)"]
    ws1.append(summary_headers)
    style_header_row(ws1, 1)

    methods = {
        "วิธีสี่เหลี่ยม (Hs_rect)": [r["Hs_rect"] for r in results],
        "วิธีสเปกตรัม (Hs_spec)": [r["Hs_spec"] for r in results],
        "วิธีข้ามศูนย์ (Hs_zc)": [r["Hs_zc"] for r in results],
    }
    for method_name, values in methods.items():
        vals = [v for v in values if v is not None and not np.isnan(v)]
        if vals:
            ws1.append([
                method_name,
                round(float(np.mean(vals)), 4),
                round(float(np.max(vals)), 4),
                round(float(np.min(vals)), 4),
                round(float(np.std(vals)), 4),
            ])
        else:
            ws1.append([method_name, "-", "-", "-", "-"])
    auto_width(ws1)

    # ── Sheet 2: Time Series ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("อนุกรมเวลา")
    ts_headers = [
        "epoch_start", "epoch_end",
        "Hs_rect (m)", "Hs_spec (m)", "Hs_zc (m)",
        "Tp (s)", "Tm02 (s)", "H_max (m)", "T_mean (s)", "wave_count",
    ]
    ws2.append(ts_headers)
    style_header_row(ws2, 1)

    for r in results:
        def fmt(v):
            if v is None or (isinstance(v, float) and np.isnan(v)):
                return ""
            if isinstance(v, float):
                return round(v, 4)
            return v

        ws2.append([
            fmt(r.get("epoch_start")),
            fmt(r.get("epoch_end")),
            fmt(r.get("Hs_rect")),
            fmt(r.get("Hs_spec")),
            fmt(r.get("Hs_zc")),
            fmt(r.get("Tp")),
            fmt(r.get("Tm02")),
            fmt(r.get("H_max")),
            fmt(r.get("T_mean")),
            r.get("wave_count", 0),
        ])
    auto_width(ws2)

    # ── Sheet 3: Spectrum ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("สเปกตรัม")
    ws3.append(["ความถี่ (Hz)", "ความหนาแน่นสเปกตรัม (m²/Hz)"])
    style_header_row(ws3, 1)

    if 0 <= selected_epoch_idx < len(results):
        sel = results[selected_epoch_idx]
        f_arr = sel.get("f_arr")
        Pxx_arr = sel.get("Pxx_arr")
        if f_arr is not None and Pxx_arr is not None:
            for f_val, pxx_val in zip(f_arr, Pxx_arr):
                ws3.append([round(float(f_val), 6), round(float(pxx_val), 8)])
    auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── PDF ──────────────────────────────────────────────────────────────────────

THAI_FONT_PATH = "/usr/share/fonts/truetype/noto/NotoSansThai-Regular.ttf"
THAI_FONT_NAME = "NotoThai"

def _register_thai_font():
    """Register NotoSansThai with ReportLab (idempotent)."""
    global THAI_FONT_NAME
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    try:
        pdfmetrics.getFont(THAI_FONT_NAME)
    except Exception:
        if os.path.exists(THAI_FONT_PATH):
            pdfmetrics.registerFont(TTFont(THAI_FONT_NAME, THAI_FONT_PATH))
        else:
            # Fallback: use Helvetica
            THAI_FONT_NAME = "Helvetica"


def _fig_to_image_rl(fig) -> "Image":
    """Convert a matplotlib Figure to a ReportLab Image object."""
    from reportlab.platypus import Image as RLImage
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, bbox_inches="tight",
                facecolor="white")
    buf.seek(0)
    return RLImage(buf, width=450, height=270)


def _make_hs_timeseries_fig(results: List[Dict[str, Any]]) -> plt.Figure:
    epochs = list(range(len(results)))
    Hs_rect = [r.get("Hs_rect", np.nan) for r in results]
    Hs_spec = [r.get("Hs_spec", np.nan) for r in results]
    Hs_zc = [r.get("Hs_zc", np.nan) for r in results]

    fig, ax = plt.subplots(figsize=(9, 3.5), facecolor="white")
    ax.set_facecolor("white")
    ax.plot(epochs, Hs_rect, "b-o", markersize=3, label="Hs_rect")
    ax.plot(epochs, Hs_spec, "r-s", markersize=3, label="Hs_spec")
    ax.plot(epochs, Hs_zc, "g-^", markersize=3, label="Hs_zc")
    ax.set_xlabel("Epoch")
    ax.set_ylabel("Hs (m)")
    ax.set_title("อนุกรมเวลาความสูงคลื่นนัยสำคัญ")
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return fig


def _make_spectrum_fig(results: List[Dict[str, Any]],
                        epoch_idx: int) -> plt.Figure:
    fig, ax = plt.subplots(figsize=(9, 3.5), facecolor="white")
    ax.set_facecolor("white")
    if 0 <= epoch_idx < len(results):
        r = results[epoch_idx]
        f_arr = r.get("f_arr")
        Pxx_arr = r.get("Pxx_arr")
        if f_arr is not None and Pxx_arr is not None:
            ax.fill_between(f_arr, Pxx_arr, alpha=0.4, color="steelblue")
            ax.plot(f_arr, Pxx_arr, color="steelblue")
            fp = r.get("fp", np.nan)
            if not np.isnan(fp):
                ax.axvline(fp, color="red", linestyle="--",
                           label=f"fp = {fp:.3f} Hz")
            ax.set_xlabel("ความถี่ (Hz)")
            ax.set_ylabel("ความหนาแน่นสเปกตรัม (m²/Hz)")
            ax.set_title(f"สเปกตรัมพลังงาน - Epoch {epoch_idx + 1}")
            ax.legend()
            ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return fig


def _make_scatter_fig(results: List[Dict[str, Any]]) -> plt.Figure:
    from scipy.stats import pearsonr

    Hs_rect = np.array([r.get("Hs_rect", np.nan) for r in results])
    Hs_spec = np.array([r.get("Hs_spec", np.nan) for r in results])
    Hs_zc = np.array([r.get("Hs_zc", np.nan) for r in results])

    fig, axes = plt.subplots(1, 2, figsize=(9, 4), facecolor="white")
    for ax in axes:
        ax.set_facecolor("white")

    for ax, y, ylabel, color in [
        (axes[0], Hs_spec, "Hs_spec (m)", "red"),
        (axes[1], Hs_zc, "Hs_zc (m)", "green"),
    ]:
        mask = np.isfinite(Hs_rect) & np.isfinite(y)
        if np.sum(mask) > 1:
            x_m, y_m = Hs_rect[mask], y[mask]
            r_val, _ = pearsonr(x_m, y_m)
            rmse = float(np.sqrt(np.mean((x_m - y_m) ** 2)))
            ax.scatter(x_m, y_m, c=color, alpha=0.6, s=20)
            lims = [min(x_m.min(), y_m.min()), max(x_m.max(), y_m.max())]
            ax.plot(lims, lims, "k--", lw=1, label="1:1")
            ax.set_xlabel("Hs_rect (m)")
            ax.set_ylabel(ylabel)
            ax.set_title(f"R²={r_val**2:.3f}  RMSE={rmse:.3f} m")
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)
    fig.tight_layout()
    return fig


def build_pdf(results: List[Dict[str, Any]],
              epoch_idx: int = 0,
              date_range: str = "",
              n_samples: int = 0) -> bytes:
    """
    Build a multi-page PDF report.
    Returns bytes.
    """
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    _register_thai_font()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    thai_style = ParagraphStyle(
        "Thai", fontName=THAI_FONT_NAME, fontSize=11, leading=16
    )
    title_style = ParagraphStyle(
        "ThaiTitle", fontName=THAI_FONT_NAME, fontSize=18, leading=24,
        alignment=1, spaceAfter=12
    )
    h2_style = ParagraphStyle(
        "ThaiH2", fontName=THAI_FONT_NAME, fontSize=14, leading=20,
        spaceBefore=12, spaceAfter=6
    )

    tbl_header_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), THAI_FONT_NAME),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#E8F0FE"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])

    story = []

    # ── Page 1: Title & Summary ────────────────────────────────────────────
    story.append(Paragraph("รายงานการวิเคราะห์ความสูงคลื่นนัยสำคัญ", title_style))
    story.append(Paragraph("Significant Wave Height (Hs) Analysis Report", title_style))
    story.append(Spacer(1, 0.5*cm))

    info_data = [
        ["รายการ", "ข้อมูล"],
        ["ช่วงเวลาข้อมูล", date_range or "-"],
        ["จำนวนตัวอย่างทั้งหมด", str(n_samples)],
        ["จำนวน Epoch ที่วิเคราะห์", str(len(results))],
    ]
    tbl_info = Table(info_data, colWidths=[8*cm, 9*cm])
    tbl_info.setStyle(tbl_header_style)
    story.append(tbl_info)
    story.append(Spacer(1, 0.5*cm))

    # Stats summary table
    story.append(Paragraph("ตารางสรุปสถิติ", h2_style))
    methods_map = {
        "วิธีสี่เหลี่ยม (Hs_rect)": [r.get("Hs_rect", np.nan) for r in results],
        "วิธีสเปกตรัม (Hs_spec)": [r.get("Hs_spec", np.nan) for r in results],
        "วิธีข้ามศูนย์ (Hs_zc)": [r.get("Hs_zc", np.nan) for r in results],
    }
    stat_rows = [["วิธีการ", "ค่าเฉลี่ย (m)", "ค่าสูงสุด (m)",
                   "ค่าต่ำสุด (m)", "S.D. (m)"]]
    for mname, vals in methods_map.items():
        v = np.array([x for x in vals if x is not None and not np.isnan(x)])
        if len(v):
            stat_rows.append([mname,
                               f"{np.mean(v):.3f}", f"{np.max(v):.3f}",
                               f"{np.min(v):.3f}", f"{np.std(v):.3f}"])
        else:
            stat_rows.append([mname, "-", "-", "-", "-"])
    tbl_stat = Table(stat_rows, colWidths=[7*cm, 3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
    tbl_stat.setStyle(tbl_header_style)
    story.append(tbl_stat)
    story.append(PageBreak())

    # ── Page 2: Hs Time Series ─────────────────────────────────────────────
    story.append(Paragraph("อนุกรมเวลาความสูงคลื่นนัยสำคัญ", h2_style))
    fig_ts = _make_hs_timeseries_fig(results)
    story.append(_fig_to_image_rl(fig_ts))
    plt.close(fig_ts)
    story.append(PageBreak())

    # ── Page 3: Spectrum ───────────────────────────────────────────────────
    story.append(Paragraph(f"สเปกตรัมพลังงาน (Epoch {epoch_idx + 1})", h2_style))
    fig_sp = _make_spectrum_fig(results, epoch_idx)
    story.append(_fig_to_image_rl(fig_sp))
    plt.close(fig_sp)
    if epoch_idx < len(results):
        r = results[epoch_idx]
        spec_info = [
            ["พารามิเตอร์", "ค่า"],
            ["Hs_spec (m)", f"{r.get('Hs_spec', np.nan):.3f}"],
            ["Tp (s)", f"{r.get('Tp', np.nan):.2f}"],
            ["Tm02 (s)", f"{r.get('Tm02', np.nan):.2f}"],
            ["fp (Hz)", f"{r.get('fp', np.nan):.4f}"],
        ]
        tbl_sp = Table(spec_info, colWidths=[8*cm, 9*cm])
        tbl_sp.setStyle(tbl_header_style)
        story.append(Spacer(1, 0.4*cm))
        story.append(tbl_sp)
    story.append(PageBreak())

    # ── Page 4: Method Comparison ──────────────────────────────────────────
    story.append(Paragraph("การเปรียบเทียบวิธีการคำนวณ", h2_style))
    fig_sc = _make_scatter_fig(results)
    story.append(_fig_to_image_rl(fig_sc))
    plt.close(fig_sc)
    story.append(PageBreak())

    # ── Page 5: Full Epoch Table ───────────────────────────────────────────
    story.append(Paragraph("ตารางข้อมูล Epoch ทั้งหมด", h2_style))
    table_headers = ["Epoch", "Hs_rect", "Hs_spec", "Hs_zc",
                      "Tp (s)", "Tm02 (s)", "H_max", "คลื่น"]
    rows = [table_headers]
    for i, r in enumerate(results):
        def fv(v, decimals=3):
            if v is None or (isinstance(v, float) and np.isnan(v)):
                return "-"
            return f"{v:.{decimals}f}"
        rows.append([
            str(i + 1),
            fv(r.get("Hs_rect")),
            fv(r.get("Hs_spec")),
            fv(r.get("Hs_zc")),
            fv(r.get("Tp"), 2),
            fv(r.get("Tm02"), 2),
            fv(r.get("H_max")),
            str(r.get("wave_count", 0)),
        ])
    col_w = [1.5*cm, 2.5*cm, 2.5*cm, 2.5*cm,
              2.5*cm, 2.5*cm, 2.5*cm, 2*cm]
    tbl_epoch = Table(rows, colWidths=col_w, repeatRows=1)
    tbl_epoch.setStyle(tbl_header_style)
    story.append(tbl_epoch)

    doc.build(story)
    return buf.getvalue()
